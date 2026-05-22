"""
Excel import/export for Test Run steps.
Ported from V3 monolith excel_service.py with minor adaptations for V4 UUID IDs.
"""
from __future__ import annotations
import io
import logging

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("excel_service")

COLUMNS = [
    ("Script Name",      30),
    ("Step No",           8),
    ("Step Description", 40),
    ("Action",           25),
    ("Input Type",       15),
    ("Field Label",      25),
    ("Input Value",      35),
    ("Screenshot",       12),
    ("Active",           10),
    ("Notes",            30),
]
COL_IDX = {name: i + 1 for i, (name, _) in enumerate(COLUMNS)}

VALID_ACTIONS = [
    "Login into Application(OJ)", "Navigate", "Click Button", "Click Link", "Click",
    "Enter Value - Text Field", "Enter Value - Dropdown", "Enter Value Text Field(Oj)",
    "Open Dropdown", "Select Option", "Dropdown Values",
    "Date Picker", "Select Date", "Fill Date",
    "Key - Enter", "Key - Tab", "Key - Press",
    "Check", "Uncheck", "Action",
]

VALID_INPUT_TYPES = [
    "Textbox", "Dropdown", "Button", "Link", "Checkbox", "Date", "Navigate", "Other", "",
]


# ── Export ──────────────────────────────────────────────────────────────────────

def build_excel(scripts_data: list[dict]) -> bytes:
    """
    Build an XLSX workbook from test run step data.

    scripts_data = [
        {
            "script_name": "Create Purchase Order",
            "steps": [
                {
                    "step_no": 1, "step_description": "Login",
                    "action": "Login into Application(OJ)",
                    "input_parameter": "Login", "input_type": "Other",
                    "default_value": "", "take_screenshot": True, "is_active": True,
                },
                ...
            ]
        },
        ...
    ]
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Data"

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border       = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    for col_name, width in COLUMNS:
        c = ws.cell(row=1, column=COL_IDX[col_name], value=col_name)
        c.font      = header_font
        c.fill      = header_fill
        c.alignment = header_align
        c.border    = border
        ws.column_dimensions[get_column_letter(COL_IDX[col_name])].width = width

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    row_no    = 2
    even_fill = PatternFill("solid", fgColor="F0F7FF")
    data_align = Alignment(vertical="center", wrap_text=False)

    for script in scripts_data:
        script_name = script.get("script_name", "")
        steps       = script.get("steps", [])

        for step in steps:
            fill = even_fill if row_no % 2 == 0 else None

            def _cell(col: str, val) -> None:
                c = ws.cell(row=row_no, column=COL_IDX[col], value=val)
                c.border    = border
                c.alignment = data_align
                if fill:
                    c.fill = fill

            _cell("Script Name",      script_name)
            _cell("Step No",          step.get("step_no", row_no - 1))
            _cell("Step Description", step.get("step_description", ""))
            _cell("Action",           step.get("action", ""))
            _cell("Input Type",       step.get("input_type", "") or "")
            _cell("Field Label",      step.get("input_parameter", "") or "")
            _cell("Input Value",      step.get("default_value", "") or "")
            _cell("Screenshot",       "Y" if step.get("take_screenshot", True) else "N")
            _cell("Active",           "Y" if step.get("is_active", True) else "N")
            _cell("Notes",            "")
            row_no += 1

        if steps:
            row_no += 1   # blank separator

    _add_hints_sheet(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _add_hints_sheet(wb: openpyxl.Workbook) -> None:
    ref = wb.create_sheet("Reference")
    ref.sheet_state = "hidden"
    ref["A1"] = "Valid Actions"
    for i, a in enumerate(VALID_ACTIONS, 2):
        ref[f"A{i}"] = a
    ref["B1"] = "Valid Input Types"
    for i, t in enumerate(VALID_INPUT_TYPES, 2):
        ref[f"B{i}"] = t


# ── Import ──────────────────────────────────────────────────────────────────────

class ImportError(Exception):
    pass


def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse uploaded XLSX file.
    Returns (rows, warnings).
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise ImportError(f"Cannot read Excel file: {exc}")

    sheet_name = "Test Data" if "Test Data" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]

    rows: list[dict]      = []
    warnings: list[str]   = []

    headers: dict[str, int] = {}
    for col in ws.iter_cols(min_row=1, max_row=1):
        for cell in col:
            if cell.value:
                headers[str(cell.value).strip()] = cell.column

    required = ["Script Name", "Step No", "Step Description", "Action"]
    missing  = [r for r in required if r not in headers]
    if missing:
        raise ImportError(f"Missing required columns: {missing}")

    def _val(row_cells, col_name: str) -> str:
        idx  = headers.get(col_name)
        if not idx:
            return ""
        cell = row_cells[idx - 1] if idx <= len(row_cells) else None
        v = cell.value if cell else None
        return str(v).strip() if v is not None else ""

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        row_cells  = list(row)
        script_val = _val(row_cells, "Script Name")
        if not script_val:
            continue

        step_no_raw   = _val(row_cells, "Step No")
        step_desc     = _val(row_cells, "Step Description")
        action        = _val(row_cells, "Action")
        input_type    = _val(row_cells, "Input Type")
        field_label   = _val(row_cells, "Field Label")
        input_value   = _val(row_cells, "Input Value")
        screenshot    = _val(row_cells, "Screenshot").upper()
        active        = _val(row_cells, "Active").upper()

        row_warnings: list[str] = []

        try:
            step_no = int(float(step_no_raw)) if step_no_raw else 0
        except ValueError:
            step_no = 0
            row_warnings.append(f"Row {row_idx}: invalid Step No '{step_no_raw}' — set to 0")

        if action and action not in VALID_ACTIONS:
            row_warnings.append(f"Row {row_idx}: unknown action '{action}' — will use as-is")

        if input_type and input_type not in VALID_INPUT_TYPES:
            row_warnings.append(f"Row {row_idx}: unknown input type '{input_type}' — set to 'Other'")
            input_type = "Other"

        warnings.extend(row_warnings)

        rows.append({
            "script_name":      script_val,
            "step_no":          step_no,
            "step_description": step_desc,
            "action":           action,
            "input_type":       input_type,
            "input_parameter":  field_label,
            "input_value":      input_value,
            "take_screenshot":  screenshot != "N",
            "is_active":        active != "N",
            "is_manual":        True,
            "_warnings":        row_warnings,
        })

    return rows, warnings